"""
Excel Live Updater module for continuous stock price updates in Excel files
This module handles real-time updates of stock prices directly in Excel spreadsheets
"""
import openpyxl
import pandas as pd
import threading
import time
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Callable
import queue
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .data_fetcher import IndianStockDataFetcher

logger = logging.getLogger(__name__)

class ExcelLiveUpdater:
    """
    Live Excel updater that continuously updates stock prices in Excel files
    """
    
    def __init__(self, excel_file_path: str, update_interval_seconds: int = 30):
        """
        Initialize the Excel live updater
        
        Arguments:
            excel_file_path: Path to the Excel file to update
            update_interval_seconds: How often to update prices (in seconds)
        """
        self.excel_file_path = Path(excel_file_path)
        self.update_interval = update_interval_seconds
        self.data_fetcher = IndianStockDataFetcher()
        
        # Threading control
        self.is_running = False
        self.update_thread = None
        self.stop_event = threading.Event()
        
        # Stock tracking
        self.tracked_stocks = []
        self.stock_rows = {}  # Map stock symbol to row number
        self.last_update_time = None
        
        # Excel workbook and worksheet references
        self.workbook = None
        self.worksheet = None
        
        # Status callback for GUI updates
        self.status_callback = None
        
        # Initialize Excel file structure
        self._initialize_excel_file()
        
    def set_status_callback(self, callback: Callable[[str], None]):
        """Set callback function for status updates"""
        self.status_callback = callback
        
    def _update_status(self, message: str):
        """Update status via callback or log"""
        if self.status_callback:
            self.status_callback(message)
        else:
            logger.info(message)
    
    def _initialize_excel_file(self):
        """Initialize or load the Excel file structure"""
        try:
            if self.excel_file_path.exists():
                # Load existing file
                self.workbook = openpyxl.load_workbook(self.excel_file_path)
                if "Live Stock Data" in self.workbook.sheetnames:
                    self.worksheet = self.workbook["Live Stock Data"]
                    self._load_existing_stocks()
                else:
                    # Create the live data sheet
                    self.worksheet = self.workbook.create_sheet("Live Stock Data")
                    self._setup_headers()
            else:
                # Create new file
                self.workbook = openpyxl.Workbook()
                self.worksheet = self.workbook.active
                self.worksheet.title = "Live Stock Data"
                self._setup_headers()
                self.workbook.save(self.excel_file_path)
                
            self._update_status(f"Excel file initialized: {self.excel_file_path}")
            
        except Exception as e:
            logger.error(f"Error initializing Excel file: {str(e)}")
            raise
    
    def _setup_headers(self):
        """Setup column headers for the live data sheet"""
        headers = [
            "Symbol", "Company Name", "Current Price (₹)", "Previous Close (₹)", 
            "Change (₹)", "Change %", "Volume", "Market Cap", "P/E Ratio", 
            "52W High (₹)", "52W Low (₹)", "Last Updated", "Source"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            
            # Apply header formatting
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
        
        # Set column widths
        column_widths = [15, 35, 15, 15, 12, 10, 15, 15, 10, 15, 15, 20, 15]
        for col, width in enumerate(column_widths, 1):
            self.worksheet.column_dimensions[chr(64 + col)].width = width
    
    def _load_existing_stocks(self):
        """Load existing stocks from the Excel file"""
        try:
            self.tracked_stocks = []
            self.stock_rows = {}
            
            # Find stocks in the worksheet
            for row in range(2, self.worksheet.max_row + 1):
                symbol_cell = self.worksheet.cell(row=row, column=1)
                if symbol_cell.value:
                    symbol = str(symbol_cell.value).strip().upper()
                    if symbol:
                        self.tracked_stocks.append(symbol)
                        self.stock_rows[symbol] = row
            
            self._update_status(f"Loaded {len(self.tracked_stocks)} existing stocks from Excel")
            
        except Exception as e:
            logger.error(f"Error loading existing stocks: {str(e)}")
    
    def add_stock(self, symbol: str) -> bool:
        """
        Add a stock to live tracking
        
        Arguments:
            symbol: Stock symbol to add
            
        Returns:
            True if successful, False otherwise
        """
        try:
            symbol = symbol.upper().strip()
            
            # Auto-add .NS if not present
            if not symbol.endswith('.NS') and not symbol.endswith('.BO'):
                symbol += '.NS'
            
            if symbol in self.tracked_stocks:
                self._update_status(f"Stock {symbol} already being tracked")
                return False
            
            # Validate symbol first
            self._update_status(f"Validating stock symbol {symbol}...")
            stock_data = self.data_fetcher.get_stock_info(symbol)
            
            if not stock_data:
                self._update_status(f"Invalid stock symbol: {symbol}")
                return False
            
            # Add to tracking list
            self.tracked_stocks.append(symbol)
            
            # Find next available row
            next_row = len(self.tracked_stocks) + 1
            self.stock_rows[symbol] = next_row
            
            # Add initial data to Excel
            self._update_stock_row(symbol, stock_data)
            self._save_workbook()
            
            self._update_status(f"Added {symbol} to live tracking")
            return True
            
        except Exception as e:
            logger.error(f"Error adding stock {symbol}: {str(e)}")
            self._update_status(f"Error adding stock: {str(e)}")
            return False
    
    def remove_stock(self, symbol: str) -> bool:
        """
        Remove a stock from live tracking
        
        Arguments:
            symbol: Stock symbol to remove
            
        Returns:
            True if successful, False otherwise
        """
        try:
            symbol = symbol.upper().strip()
            
            if symbol not in self.tracked_stocks:
                return False
            
            # Remove from tracking
            self.tracked_stocks.remove(symbol)
            row_to_remove = self.stock_rows.pop(symbol, None)
            
            if row_to_remove:
                # Delete the row in Excel
                self.worksheet.delete_rows(row_to_remove)
                
                # Update row mappings for remaining stocks
                for stock, row in self.stock_rows.items():
                    if row > row_to_remove:
                        self.stock_rows[stock] = row - 1
                
                self._save_workbook()
            
            self._update_status(f"Removed {symbol} from live tracking")
            return True
            
        except Exception as e:
            logger.error(f"Error removing stock {symbol}: {str(e)}")
            return False
    
    def _update_stock_row(self, symbol: str, stock_data: Dict):
        """Update a single stock row in Excel with new data"""
        try:
            if symbol not in self.stock_rows:
                return
            
            row = self.stock_rows[symbol]
            
            # Prepare values
            values = [
                symbol,
                stock_data.get('name', symbol),
                stock_data.get('current_price', 0),
                stock_data.get('previous_close', 0),
                stock_data.get('change', 0),
                stock_data.get('change_percent', 0),
                stock_data.get('volume', 0),
                stock_data.get('market_cap', ''),
                stock_data.get('pe_ratio', ''),
                stock_data.get('fifty_two_week_high', 0),
                stock_data.get('fifty_two_week_low', 0),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                stock_data.get('source', 'Unknown')
            ]
            
            # Update cells
            for col, value in enumerate(values, 1):
                cell = self.worksheet.cell(row=row, column=col)
                cell.value = value
                
                # Apply formatting based on column type
                if col in [3, 4, 5, 10, 11]:  # Price columns
                    cell.number_format = '₹#,##0.00'
                elif col == 6:  # Change percentage
                    cell.number_format = '0.00%'
                    # Apply conditional formatting for gains/losses
                    if isinstance(value, (int, float)) and value != 0:
                        if value > 0:
                            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                elif col == 5:  # Change amount
                    cell.number_format = '₹#,##0.00'
                    # Apply conditional formatting
                    if isinstance(value, (int, float)) and value != 0:
                        if value > 0:
                            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                elif col == 7:  # Volume
                    cell.number_format = '#,##0'
                elif col == 9:  # P/E ratio
                    cell.number_format = '0.00'
                
                # Add border
                cell.border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin")
                )
                
                # Center alignment
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
        except Exception as e:
            logger.error(f"Error updating row for {symbol}: {str(e)}")
    
    def _save_workbook(self):
        """Save the workbook to file"""
        try:
            self.workbook.save(self.excel_file_path)
        except PermissionError:
            # File might be open in Excel
            logger.warning(f"Could not save {self.excel_file_path} - file may be open in Excel")
        except Exception as e:
            logger.error(f"Error saving workbook: {str(e)}")
    
    def start_live_updates(self):
        """Start the live update process"""
        if self.is_running:
            self._update_status("Live updates already running")
            return
        
        if not self.tracked_stocks:
            self._update_status("No stocks to track - add stocks first")
            return
        
        self.is_running = True
        self.stop_event.clear()
        
        def update_loop():
            self._update_status(f"Started live Excel updates (every {self.update_interval}s)")
            
            while not self.stop_event.is_set():
                try:
                    # Update all tracked stocks
                    self._update_all_stocks()
                    
                    # Wait for next update or stop signal
                    if self.stop_event.wait(self.update_interval):
                        break
                        
                except Exception as e:
                    logger.error(f"Error in update loop: {str(e)}")
                    time.sleep(5)  # Wait before retrying
            
            self.is_running = False
            self._update_status("Live Excel updates stopped")
        
        self.update_thread = threading.Thread(target=update_loop, daemon=True)
        self.update_thread.start()
    
    def stop_live_updates(self):
        """Stop the live update process"""
        if not self.is_running:
            return
        
        self.stop_event.set()
        if self.update_thread and self.update_thread.is_alive():
            self.update_thread.join(timeout=2)
        
        self.is_running = False
        self._update_status("Stopped live Excel updates")
    
    def _update_all_stocks(self):
        """Update all tracked stocks in Excel"""
        try:
            if not self.tracked_stocks:
                return
            
            self._update_status(f"Updating {len(self.tracked_stocks)} stocks in Excel...")
            
            # Fetch data for all stocks
            stock_data_df = self.data_fetcher.get_multiple_stocks_batch(
                self.tracked_stocks, 
                force_refresh=True
            )
            
            if stock_data_df.empty:
                self._update_status("No data received for stocks")
                return
            
            # Update each stock in Excel
            updated_count = 0
            for _, row in stock_data_df.iterrows():
                symbol = row.get('symbol')
                if symbol in self.stock_rows:
                    self._update_stock_row(symbol, row.to_dict())
                    updated_count += 1
            
            # Save the workbook
            self._save_workbook()
            
            self.last_update_time = datetime.now()
            self._update_status(f"Updated {updated_count} stocks in Excel at {self.last_update_time.strftime('%H:%M:%S')}")
            
        except Exception as e:
            logger.error(f"Error updating all stocks: {str(e)}")
            self._update_status(f"Update failed: {str(e)}")
    
    def get_status(self) -> Dict:
        """Get current status of the live updater"""
        return {
            'is_running': self.is_running,
            'tracked_stocks_count': len(self.tracked_stocks),
            'tracked_stocks': self.tracked_stocks.copy(),
            'excel_file': str(self.excel_file_path),
            'update_interval': self.update_interval,
            'last_update': self.last_update_time.isoformat() if self.last_update_time else None
        }
    
    def set_update_interval(self, interval_seconds: int):
        """Change the update interval"""
        self.update_interval = max(10, interval_seconds)  # Minimum 10 seconds
        self._update_status(f"Update interval changed to {self.update_interval} seconds")
    
    def manual_update(self):
        """Trigger a manual update of all stocks"""
        if self.is_running:
            # Don't interfere with automatic updates
            self._update_status("Manual update skipped - automatic updates running")
            return
        
        try:
            self._update_all_stocks()
        except Exception as e:
            self._update_status(f"Manual update failed: {str(e)}")
    
    def cleanup(self):
        """Clean up resources"""
        self.stop_live_updates()
        if self.workbook:
            try:
                self.workbook.close()
            except:
                pass
