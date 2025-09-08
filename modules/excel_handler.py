"""
Excel handler module for saving and managing Indian stock data in Excel format
Professional formatting with Indian currency and market-specific styling
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from pathlib import Path
import logging
from typing import Optional, Dict, List
import shutil

logger = logging.getLogger(__name__)

class IndianStockExcelHandler:
    """
    Professional Excel handler for Indian stock market data
    """
    
    def __init__(self, data_dir: Path):
        """
        Initialize Excel handler
        Arguments:
            data_dir: Directory to save Excel files
        """
        self.data_dir = Path(data_dir)
        self.data_dir.mkdir(exist_ok=True)
        
    def save_stock_data(self, df: pd.DataFrame, filename: Optional[str] = None) -> str:
        """
        Save Indian stock data to Excel file with professional formatting
        Arguments:
            df: DataFrame containing Indian stock data
            filename: Optional custom filename
            
        Returns:
            Path to the saved file
        """
        if df.empty:
            raise ValueError("DataFrame is empty, cannot save to Excel")
        
        # Generate filename if not provided
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"indian_stocks_{timestamp}.xlsx"
        
        filepath = self.data_dir / filename
        
        try:
            # Create a copy of the dataframe for processing
            df_excel = df.copy()
            
            # Format numeric columns
            numeric_columns = ['current_price', 'previous_close', 'change', 'change_percent', 
                             'volume', 'market_cap', 'pe_ratio', 'dividend_yield',
                             'fifty_two_week_high', 'fifty_two_week_low']
            
            for col in numeric_columns:
                if col in df_excel.columns:
                    df_excel[col] = pd.to_numeric(df_excel[col], errors='coerce')
            
            # Format timestamp column
            if 'timestamp' in df_excel.columns:
                df_excel['timestamp'] = pd.to_datetime(df_excel['timestamp']).dt.strftime('%Y-%m-%d %H:%M:%S')
            
            # Save to Excel with professional formatting
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df_excel.to_excel(writer, sheet_name='Indian Stocks', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Indian Stocks']
                
                # Apply Indian market specific formatting
                self._apply_indian_formatting(worksheet, df_excel)
                
                # Add summary sheet
                self._create_summary_sheet(workbook, df_excel)
            
            logger.info(f"Indian stock data saved to {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error saving Excel file: {str(e)}")
            raise
    
    def _apply_indian_formatting(self, worksheet, df: pd.DataFrame):
        """Apply professional formatting for Indian stock data"""
        
        # Define Indian market specific styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Indian blue
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        cell_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Format headers
        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            
            # Auto-adjust column width for Indian stock data
            max_length = max(len(str(column)), 12)
            if column in ['name']:
                max_length = 30  # Longer for Indian company names
            elif column in ['symbol']:
                max_length = 15  # Accommodate .NS/.BO suffixes
            elif column in ['current_price', 'previous_close', 'change']:
                max_length = 15  # Indian rupee values
            elif column in ['market_cap']:
                max_length = 18  # Larger Indian companies
            elif column in ['source']:
                max_length = 12
            
            worksheet.column_dimensions[cell.column_letter].width = max_length
        
        # Format data cells with Indian market specific formatting
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = cell_alignment
                cell.border = border
                
                # Apply conditional formatting for change columns
                if df.columns[col_num - 1] == 'change':
                    if cell.value and float(cell.value) > 0:
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Green for gains
                    elif cell.value and float(cell.value) < 0:
                        cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Orange for losses
                
                elif df.columns[col_num - 1] == 'change_percent':
                    if cell.value and float(cell.value) > 0:
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    elif cell.value and float(cell.value) < 0:
                        cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        # Apply Indian rupee number formatting
        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            for cell in row:
                col_name = df.columns[cell.column - 1]
                
                if col_name in ['current_price', 'previous_close', 'change', 'fifty_two_week_high', 'fifty_two_week_low']:
                    cell.number_format = '₹#,##0.00'
                elif col_name in ['change_percent', 'dividend_yield']:
                    cell.number_format = '0.00%'
                elif col_name in ['volume']:
                    cell.number_format = '#,##0'
                elif col_name in ['market_cap']:
                    cell.number_format = '₹#,##0,,'  # In crores
                elif col_name in ['pe_ratio']:
                    cell.number_format = '0.00'
    
    def _create_summary_sheet(self, workbook, df: pd.DataFrame):
        """Create a summary sheet with Indian market statistics"""
        
        summary_sheet = workbook.create_sheet(title="Market Summary")
        
        # Calculate Indian market specific statistics
        stats = {
            'Total Indian Stocks': len(df),
            'Average Price (₹)': f"₹{df['current_price'].mean():.2f}" if 'current_price' in df.columns else 'N/A',
            'Gainers': len(df[df['change'] > 0]) if 'change' in df.columns else 0,
            'Losers': len(df[df['change'] < 0]) if 'change' in df.columns else 0,
            'Unchanged': len(df[df['change'] == 0]) if 'change' in df.columns else 0,
            'Total Volume': f"{df['volume'].sum():,}" if 'volume' in df.columns else 'N/A',
            'Highest Gainer (%)': f"{df['change_percent'].max():.2f}%" if 'change_percent' in df.columns else 'N/A',
            'Highest Loser (%)': f"{df['change_percent'].min():.2f}%" if 'change_percent' in df.columns else 'N/A',
            'Generated': datetime.now().strftime('%d/%m/%Y %H:%M:%S IST')
        }
        
        # Write summary data with Indian styling
        summary_sheet['A1'] = 'Indian Stock Market Summary'
        summary_sheet['A1'].font = Font(bold=True, size=16, color="1F4E79")
        
        row = 3
        for key, value in stats.items():
            summary_sheet[f'A{row}'] = key
            summary_sheet[f'B{row}'] = value
            summary_sheet[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Auto-adjust column widths
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 20
    
    def load_stock_data(self, filepath: str) -> pd.DataFrame:
        """
        Load Indian stock data from Excel file
        Arguments:
            filepath: Path to Excel file
        Returns:
            DataFrame containing the loaded data
        """
        try:
            filepath = Path(filepath)
            if not filepath.exists():
                raise FileNotFoundError(f"File {filepath} does not exist")
            
            df = pd.read_excel(filepath, sheet_name='Indian Stocks')
            logger.info(f"Loaded Indian stock data from {filepath}")
            return df
            
        except Exception as e:
            logger.error(f"Error loading Excel file: {str(e)}")
            raise
    
    def get_saved_files(self) -> List[Dict]:
        """
        Get list of saved Excel files with metadata
        Returns:
            List of dictionaries containing file information
        """
        files = []
        
        for filepath in self.data_dir.glob("*.xlsx"):
            try:
                stat = filepath.stat()
                files.append({
                    'filename': filepath.name,
                    'filepath': str(filepath),
                    'size': stat.st_size,
                    'modified': datetime.fromtimestamp(stat.st_mtime),
                    'size_mb': round(stat.st_size / (1024 * 1024), 2)
                })
            except Exception as e:
                logger.warning(f"Error reading file info for {filepath}: {str(e)}")
        
        # Sort by modification time (newest first)
        files.sort(key=lambda x: x['modified'], reverse=True)
        return files
    
    def delete_file(self, filepath: str) -> bool:
        """
        Delete an Excel file
        Arguments:
            filepath: Path to file to delete
        Returns:
            True if successful, False otherwise
        """
        try:
            filepath = Path(filepath)
            if filepath.exists():
                filepath.unlink()
                logger.info(f"Deleted file {filepath}")
                return True
            return False
            
        except Exception as e:
            logger.error(f"Error deleting file: {str(e)}")
            return False
    
    def cleanup_old_files(self, keep_count: int = 10):
        """
        Clean up old Excel files, keeping only the most recent ones
        Arguments:
            keep_count: Number of recent files to keep
        """
        try:
            files = self.get_saved_files()
            
            if len(files) > keep_count:
                files_to_delete = files[keep_count:]
                
                for file_info in files_to_delete:
                    self.delete_file(file_info['filepath'])
                
                logger.info(f"Cleaned up {len(files_to_delete)} old files")
                
        except Exception as e:
            logger.error(f"Error during cleanup: {str(e)}")
